"""Step 2 — Google Maps data collection (free, no API key).

Primary path: scrape Google Maps with Playwright (same source as the paid Places API).
Pure, testable helpers (model / mock / dedup / listing-parse) are kept separate from the
browser-driving code so they can be unit-tested without a network or browser.
A deterministic `mock=True` mode makes the whole dashboard demoable offline.
"""
from __future__ import annotations

import json
import random
import re
import threading
import time
import urllib.parse
from datetime import datetime
from pathlib import Path

import requests

try:
    import config
except ModuleNotFoundError:  # pragma: no cover
    import importlib
    config = importlib.import_module("config")


# --------------------------------------------------------------------------- data model
RESULT_COLUMNS = [
    "name", "rating", "user_ratings_total", "types", "formatted_address",
    "formatted_phone_number", "website", "opening_hours", "business_status", "price_level",
    "lat", "lng", "place_url", "place_id",
    "source_query_rank", "source_query", "source_category", "result_position",
    "fetched_at", "status",
]


def empty_result_row() -> dict:
    """A result row with safe defaults: '' for strings, None for numbers."""
    return {
        "name": "", "rating": None, "user_ratings_total": None, "types": "",
        "formatted_address": "", "formatted_phone_number": None, "website": "",
        "opening_hours": "", "business_status": "OPERATIONAL", "price_level": None,
        "lat": None, "lng": None, "place_url": "", "place_id": "",
        "source_query_rank": None, "source_query": "", "source_category": "",
        "result_position": None, "fetched_at": "", "status": "OK",
    }


def dedup_key(place_url: str) -> str:
    """Stable identity for a clinic, extracted from its Google Maps URL."""
    if not place_url:
        return ""
    m = re.search(r"[?&]cid=([0-9]+)", place_url)
    if m:
        return m.group(1)
    m = re.search(r"place_id:([A-Za-z0-9_\-]+)", place_url)
    if m:
        return m.group(1)
    m = re.search(r"!1s(0x[0-9a-fA-F]+:0x[0-9a-fA-F]+)", place_url)  # feature-id form
    if m:
        return m.group(1)
    return place_url


def parse_listing(raw: dict, query_row: dict, position: int) -> dict:
    """Map a raw scraped/mock listing dict onto the canonical result-row contract."""
    row = empty_result_row()
    url = raw.get("url", "") or ""
    row.update(
        name=raw.get("name", "") or "",
        rating=raw.get("rating"),
        user_ratings_total=raw.get("reviews"),
        types=raw.get("types", "") or "",
        formatted_address=raw.get("address", "") or "",
        formatted_phone_number=raw.get("phone"),
        website=raw.get("website", "") or "",
        opening_hours=raw.get("opening_hours", "") or "",
        business_status=(raw.get("status_label") or "CLOSED") if raw.get("closed") else "OPERATIONAL",
        price_level=raw.get("price_level"),
        lat=raw.get("lat"),
        lng=raw.get("lng"),
        place_url=url,
        place_id=dedup_key(url),
        source_query_rank=query_row.get("rank"),
        source_query=query_row.get("search_query", ""),
        source_category=query_row.get("category", ""),
        result_position=position,
        fetched_at=datetime.now().isoformat(timespec="seconds"),
        status="OK",
    )
    return row


# --------------------------------------------------------------------------- mock data
_MOCK_NAMES = [
    "Sai Skin & Hair Clinic", "Guntur Derma Care", "Sri Sai Skin Clinic",
    "Lakshmi Skin Care Centre", "Apollo Skin Clinic", "Sravani Dermatology",
    "Vasavi Skin & Laser Clinic", "Sunrise Skin Hospital", "Sree Ramya Skin Clinic",
    "Amaravathi Skin Care", "Renova Skin Clinic", "Padma Dermatology Centre",
    "Care Well Skin Clinic", "Glow Derma Clinic", "Sushrutha Skin Clinic",
    "NRI Skin Institute", "Vijaya Skin Care", "KIMS Skin Department",
    "Hema Skin & Cosmetology", "Sri Krishna Skin Clinic", "Deccan Derma",
    "Pranaam Skin Clinic", "Akshara Skin Care", "Tejaswi Dermatology",
    "Manipal Skin Clinic", "Skin Solutions Guntur", "Aesthetica Skin Studio",
    "Reddy's Skin & Laser",
]


def _mock_pool() -> list[dict]:
    """Deterministic pool of fake Guntur clinics with realistic variety (no randomness)."""
    ratings = [None, 2.8, 3.2, 3.6, 4.0, 4.3, 4.6, 4.9]
    reviews = [0, 5, 12, 18, 40, 90, 150, 320]
    pool = []
    for idx, name in enumerate(_MOCK_NAMES):
        website = "" if idx % 3 == 0 else f"https://{re.sub(r'[^a-z]', '', name.lower())[:12]}.example.com"
        phone = None if idx % 4 == 0 else f"+91 9{(40000000 + idx * 137):08d}"
        pool.append({
            "name": name,
            "rating": ratings[idx % len(ratings)],
            "reviews": reviews[idx % len(reviews)],
            "types": "Dermatologist" if idx % 2 == 0 else "Skin care clinic",
            "address": f"{idx + 1}-{idx * 3 + 7}, Brodipet, Guntur, Andhra Pradesh 522002",
            "phone": phone,
            "website": website,
            "opening_hours": "Mon-Sat 10:00-20:00",
            "closed": (idx % 13 == 0 and idx != 0),
            "status_label": "TEMPORARILY_CLOSED",
            "price_level": None,
            "lat": round(16.2990 + (idx % 9) * 0.0035, 6),
            "lng": round(80.4250 + (idx % 7) * 0.0040, 6),
            "url": f"https://www.google.com/maps/place/?cid={1000 + idx}",
        })
    return pool


def make_mock_results(query_rows: list[dict], per_query: int | None = None) -> list[dict]:
    """Build deterministic result rows so clinics overlap across queries (realistic appearances)."""
    per_query = per_query or config.RESULTS_PER_QUERY
    pool = _mock_pool()
    rows: list[dict] = []
    for q in query_rows:
        start = (int(q.get("rank", 1)) * 2) % len(pool)  # shift the window per query
        for pos in range(1, per_query + 1):
            raw = pool[(start + pos - 1) % len(pool)]
            rows.append(parse_listing(raw, q, pos))
    return rows


# --------------------------------------------------------------------------- OSM fallback
def geocode_osm(address: str):
    """Free, keyless geocoding via OpenStreetMap Nominatim. Returns (lat, lng) or None."""
    if not address:
        return None
    try:
        resp = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": address, "format": "json", "limit": 1},
            headers={"User-Agent": "derma-intel/1.0 (research)"},
            timeout=15,
        )
        if resp.status_code != 200:
            return None
        data = resp.json()
        if not data:
            return None
        return (float(data[0]["lat"]), float(data[0]["lon"]))
    except Exception:
        return None


# --------------------------------------------------------------------------- cache
def _load_cache() -> dict:
    try:
        with open(config.MAPS_CACHE, "r", encoding="utf-8") as fh:
            return json.load(fh)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_cache(cache: dict) -> None:
    Path(config.MAPS_CACHE).parent.mkdir(parents=True, exist_ok=True)
    with open(config.MAPS_CACHE, "w", encoding="utf-8") as fh:
        json.dump(cache, fh, ensure_ascii=False, indent=2)


def _details_cache_path() -> str:
    return str(Path(config.CACHE_DIR) / "maps_details.json")


def _load_details_cache() -> dict:
    """Per-clinic detail cache (keyed by place id) so a clinic is enriched only once across queries."""
    try:
        with open(_details_cache_path(), "r", encoding="utf-8") as fh:
            return json.load(fh)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_details_cache(details: dict) -> None:
    Path(config.CACHE_DIR).mkdir(parents=True, exist_ok=True)
    with open(_details_cache_path(), "w", encoding="utf-8") as fh:
        json.dump(details, fh, ensure_ascii=False, indent=2)


def _polite_sleep() -> None:
    time.sleep(random.uniform(config.SCRAPER_MIN_DELAY_S, config.SCRAPER_MAX_DELAY_S))


# --------------------------------------------------------------------------- scraper
def _clean(s: str) -> str:
    """Strip Google's private-use icon glyphs (U+E000–U+F8FF) + replacement chars, collapse spaces."""
    if not s:
        return ""
    s = "".join(ch for ch in s if not (0xE000 <= ord(ch) <= 0xF8FF))
    s = s.replace("�", " ")
    return re.sub(r"\s+", " ", s).strip()


def _text_or_none(locator):
    try:
        if locator.count() == 0:
            return None
        return _clean(locator.first.inner_text(timeout=1200))
    except Exception:
        return None


def _extract_cards(page, max_results: int) -> list[dict]:
    """Extract up to `max_results` listing dicts from a loaded Google Maps results feed.

    Targets Google's current result-card structure (div.Nv2PK + rating/review spans), with an
    anchor-based fallback. Markup is unstable, so this is the part to re-tune if Google changes it.
    """
    raws: list[dict] = []
    cards = page.locator("div.Nv2PK")
    count = cards.count()
    if count == 0:  # fallback: one card per place anchor
        anchors = page.locator('a[href*="/maps/place/"]')
        count = anchors.count()
        cards = None

    for i in range(min(count, max_results)):
        card = cards.nth(i) if cards is not None else page.locator('a[href*="/maps/place/"]').nth(i)
        raw: dict = {"name": "", "url": ""}
        try:
            link = card.locator('a[href*="/maps/place/"]').first
            raw["url"] = link.get_attribute("href") or ""
            raw["name"] = _clean(link.get_attribute("aria-label") or "")
            if not raw["name"]:
                raw["name"] = _text_or_none(card.locator("div.qBF1Pd, .fontHeadlineSmall")) or ""

            # rating + reviews: the star widget's aria-label ("4.9 stars 86 Reviews") carries both
            al = ""
            try:
                imgs = card.locator('[role="img"]')
                for j in range(min(imgs.count(), 6)):
                    cand = imgs.nth(j).get_attribute("aria-label") or ""
                    if re.search(r"star", cand, re.I):
                        al = cand
                        break
            except Exception:
                pass
            if al:
                mr = re.search(r"(\d(?:[.,]\d)?)\s*star", al, re.I)
                if mr:
                    raw["rating"] = float(mr.group(1).replace(",", "."))
                mv = re.search(r"([\d,]+)\s*review", al, re.I)
                if mv:
                    raw["reviews"] = int(mv.group(1).replace(",", ""))
            # fallbacks via explicit spans
            if raw.get("rating") is None:
                rt = _text_or_none(card.locator("span.MW4etd"))
                if rt:
                    m = re.search(r"\d(?:[.,]\d)?", rt)
                    if m:
                        raw["rating"] = float(m.group(0).replace(",", "."))
            if raw.get("reviews") is None:
                rv = _text_or_none(card.locator("span.UY7F9"))
                if rv:
                    m = re.search(r"([\d,]+)", rv)
                    if m:
                        raw["reviews"] = int(m.group(1).replace(",", ""))

            # website button (only present when the place has a site)
            try:
                wl = card.locator('a[data-value="Website"]').first
                if wl.count() > 0:
                    href = wl.get_attribute("href") or ""
                    if href and "google.com" not in href:
                        raw["website"] = href
            except Exception:
                pass

            txt = _clean(_safe_inner_text(card))
            raw["closed"] = ("Permanently closed" in txt) or ("Temporarily closed" in txt)
            # phone (Indian formats) from the card info line
            mp = re.search(r"(?:\+?91[\s-]?)?0?\d{3,5}[\s-]\d{5,6}", txt)
            if mp:
                raw["phone"] = mp.group(0).strip()
            # best-effort address: a "·"-separated chunk that has a digit and a comma/pincode
            for part in re.split(r"[·•]", txt):
                part = part.strip()
                if re.search(r"\d", part) and ("," in part or re.search(r"\d{6}", part)):
                    raw["address"] = part
                    break
            # lat/lng from the place URL (!3dLAT!4dLNG)
            mll = re.search(r"!3d(-?\d+\.\d+)!4d(-?\d+\.\d+)", raw["url"])
            if mll:
                raw["lat"], raw["lng"] = float(mll.group(1)), float(mll.group(2))
        except Exception:
            pass
        raws.append(raw)
    return raws


def _safe_inner_text(locator) -> str:
    try:
        return locator.inner_text(timeout=1500)
    except Exception:
        return ""


def _enrich_from_place(page, raw: dict) -> None:
    """Open a place page and pull clean website/phone/address/status from its detail panel.

    Uses Google's stable `data-item-id` attributes (authority=website, phone:tel:=phone, address).
    Best-effort: any failure leaves the list-card values untouched.
    """
    try:
        page.goto(raw["url"], wait_until="domcontentloaded")
        page.wait_for_selector("button[data-item-id], a[data-item-id]", timeout=8000)
    except Exception:
        return
    try:
        a = page.locator('a[data-item-id="authority"]').first
        if a.count() > 0:
            href = a.get_attribute("href") or ""
            if href and "google.com" not in href:
                raw["website"] = href
    except Exception:
        pass
    try:
        b = page.locator('button[data-item-id^="phone:tel:"]').first
        if b.count() > 0:
            did = b.get_attribute("data-item-id") or ""
            raw["phone"] = did.replace("phone:tel:", "").strip() or raw.get("phone")
    except Exception:
        pass
    try:
        b = page.locator('button[data-item-id="address"]').first
        if b.count() > 0:
            label = _clean(b.get_attribute("aria-label") or "")
            raw["address"] = re.sub(r"^Address:\s*", "", label).strip()
    except Exception:
        pass
    # review count (authoritative): an aria-label like "86 reviews" in the place header
    try:
        revs = page.locator('[aria-label*="review" i]')
        for j in range(min(revs.count(), 10)):
            al = revs.nth(j).get_attribute("aria-label") or ""
            m = re.search(r"([\d,]+)\s*review", al, re.I)
            if m:
                raw["reviews"] = int(m.group(1).replace(",", ""))
                break
    except Exception:
        pass
    try:
        main = _clean(_safe_inner_text(page.locator('div[role="main"]')))
        if "Permanently closed" in main or "Temporarily closed" in main:
            raw["closed"] = True
    except Exception:
        pass


def _run_browser(query: str, max_results: int, details_cache: dict | None = None) -> list[dict]:
    """Run the sync-Playwright scrape in a worker thread.

    Streamlit's script-runner thread has a *running* asyncio loop, and Playwright's sync API refuses
    to run there ("Sync API inside the asyncio loop"). A fresh worker thread has no event loop, so the
    sync API works — while the caller's loop and any progress callbacks stay on the main thread.
    """
    box: dict = {}

    def _work():
        try:
            box["raws"] = _run_browser_impl(query, max_results, details_cache)
        except BaseException as exc:  # captured and re-raised on the caller thread
            box["err"] = exc

    t = threading.Thread(target=_work, daemon=True)
    t.start()
    t.join()
    if "err" in box:
        raise box["err"]
    return box.get("raws", [])


def _run_browser_impl(query: str, max_results: int, details_cache: dict | None = None) -> list[dict]:
    """Open Google Maps for a query and return raw listing dicts. Network/browser side-effect."""
    from playwright.sync_api import sync_playwright

    details_cache = {} if details_cache is None else details_cache
    # AI-generated queries already imply the specialty; only ensure the city is present.
    search_text = query if "guntur" in query.lower() else f"{query} Guntur"
    term = urllib.parse.quote_plus(search_text)
    url = f"https://www.google.com/maps/search/{term}"
    raws: list[dict] = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=config.SCRAPER_HEADLESS)
        context = browser.new_context(
            locale=config.SCRAPER_LOCALE, user_agent=config.SCRAPER_USER_AGENT
        )
        page = context.new_page()
        page.set_default_timeout(config.SCRAPER_PAGE_TIMEOUT_S * 1000)
        try:
            page.goto(url, wait_until="domcontentloaded")
            # consent screen
            for label in ("Reject all", "Accept all", "I agree"):
                try:
                    btn = page.get_by_role("button", name=label)
                    if btn.count() > 0:
                        btn.first.click(timeout=3000)
                        break
                except Exception:
                    pass
            try:
                page.wait_for_selector('a[href*="/maps/place/"]', timeout=12000)
            except Exception:
                pass
            # scroll the feed to lazy-load results
            feed = page.locator('[role="feed"]')
            last = 0
            for _ in range(12):
                links = page.locator('a[href*="/maps/place/"]')
                n = links.count()
                if n >= max_results or n == last:
                    if n == last:
                        break
                last = n
                try:
                    if feed.count() > 0:
                        feed.evaluate("el => el.scrollBy(0, el.scrollHeight)")
                    else:
                        page.mouse.wheel(0, 3000)
                except Exception:
                    page.mouse.wheel(0, 3000)
                time.sleep(1.2)
            raws = _extract_cards(page, max_results)
            # Phase 2: enrich each place from its detail panel (clean website/phone/address).
            # Reuse a cross-query cache so a clinic seen in an earlier query isn't re-opened.
            if config.SCRAPER_OPEN_DETAILS and raws:
                detail = context.new_page()
                detail.set_default_timeout(config.SCRAPER_PAGE_TIMEOUT_S * 1000)
                try:
                    for r in raws:
                        if not r.get("url"):
                            continue
                        key = dedup_key(r["url"])
                        cached = details_cache.get(key) if key else None
                        if cached:  # already enriched this clinic in this or a prior run
                            for field, value in cached.items():
                                if value not in (None, ""):
                                    r[field] = value
                            continue
                        _enrich_from_place(detail, r)
                        if key:
                            details_cache[key] = {
                                f: r.get(f) for f in ("website", "phone", "address", "reviews", "closed")
                            }
                        _polite_sleep()
                finally:
                    detail.close()
        finally:
            browser.close()
    return raws


def scrape_query(query_row: dict, max_results: int, cache: dict,
                 details_cache: dict | None = None) -> list[dict]:
    """Scrape one query (cache-first), map to result rows, dedupe + OSM-geocode gaps."""
    q = query_row["search_query"]
    if q in cache:
        raws = cache[q]
    else:
        raws = []
        for attempt in range(config.SCRAPER_MAX_RETRIES):
            try:
                raws = _run_browser(q, max_results, details_cache)
                if raws:
                    break
            except Exception:
                time.sleep(2 ** (attempt + 1))
        cache[q] = raws
        _polite_sleep()

    if not raws:
        fail = empty_result_row()
        fail.update(source_query_rank=query_row.get("rank"), source_query=q,
                    source_category=query_row.get("category", ""), result_position=1,
                    status="FETCH_FAILED", fetched_at=datetime.now().isoformat(timespec="seconds"))
        return [fail]

    rows, seen = [], set()
    for pos, raw in enumerate(raws[:max_results], start=1):
        row = parse_listing(raw, query_row, pos)
        key = dedup_key(row["place_url"]) or row["name"].lower()
        if config.USE_OSM_FALLBACK and (row["lat"] is None or row["lng"] is None) and row["formatted_address"]:
            geo = geocode_osm(row["formatted_address"])
            if geo:
                row["lat"], row["lng"] = geo
        if key and key in seen:
            continue
        seen.add(key)
        rows.append(row)
    return rows


def collect(query_rows: list[dict], mock: bool = False, progress_cb=None) -> list[dict]:
    """Collect result rows for all queries. `mock=True` returns deterministic offline data."""
    if mock:
        rows: list[dict] = []
        for q in query_rows:
            rows.extend(make_mock_results([q], per_query=config.RESULTS_PER_QUERY))
        return rows

    cache = _load_cache()
    details_cache = _load_details_cache()
    all_rows: list[dict] = []
    n = len(query_rows)
    for i, q in enumerate(query_rows, start=1):
        if progress_cb:
            progress_cb(i, n, q["search_query"])
        all_rows.extend(scrape_query(q, config.RESULTS_PER_QUERY, cache, details_cache))
    _save_cache(cache)
    _save_details_cache(details_cache)
    return all_rows


# --------------------------------------------------------------------------- excel
def save_results_xlsx(rows: list[dict], path: str | None = None) -> str:
    """Write result rows to a formatted .xlsx (bold frozen header, alternating row fill)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = path or config.RESULTS_XLSX
    Path(path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Maps Results"

    header_fill = PatternFill("solid", fgColor="DDEEFF")
    bold = Font(bold=True)
    headers = [c.replace("_", " ").title() for c in RESULT_COLUMNS]
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = bold
        cell.fill = header_fill

    alt_fill = PatternFill("solid", fgColor="F3F4F6")
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, key in enumerate(RESULT_COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row.get(key))
            if r_idx % 2 == 1:  # alternating shade
                cell.fill = alt_fill

    for c_idx, key in enumerate(RESULT_COLUMNS, start=1):
        longest = len(headers[c_idx - 1])
        for row in rows:
            longest = max(longest, len(str(row.get(key, ""))))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(longest + 2, 50)

    ws.freeze_panes = "A2"
    wb.save(path)
    return path
