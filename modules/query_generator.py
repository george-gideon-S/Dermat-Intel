"""Step 1 — Query generation (manual AI paste, no API key).

Workflow:
1. `build_ai_prompt()` gives the user a prompt to paste into any free AI tool.
2. The user pastes the AI's answer back; `parse_pasted_queries()` turns it into clean rows,
   auto-deriving category / user_intent / search_strength_score from each query's wording.
3. `save_queries_xlsx()` writes the formatted workbook.
"""
from __future__ import annotations

import re
from pathlib import Path

try:  # config is importable both as `config` (app) and `..config` is not needed (flat layout)
    import config
except ModuleNotFoundError:  # pragma: no cover - fallback when run oddly
    import importlib
    config = importlib.import_module("config")


# --------------------------------------------------------------------------- prompt
def build_ai_prompt() -> str:
    """Return the ready-to-copy prompt the user pastes into a free AI tool."""
    return (
        "You are a local SEO and healthcare search-behavior expert for India.\n\n"
        "Generate exactly 50 high-intent Google search queries that people in Guntur, "
        "Andhra Pradesh use before visiting a dermatologist. Base them on real doctor-search "
        "behavior: best/top ranking terms, near-me local intent, fees, reviews, appointment "
        "booking, comparison, and symptom / condition-based searches (acne, hair fall, "
        "pigmentation, eczema, psoriasis, fungal infection, etc.). Order them by estimated "
        "monthly search strength (strongest first).\n\n"
        "Return ONLY the 50 queries as a single comma-separated list of single-quoted strings. "
        "No numbering, no explanation, no markdown, no headings. Use this exact format:\n\n"
        "'best dermatologist in Guntur', 'skin specialist near me Guntur', "
        "'dermatologist fees in Guntur', 'acne treatment doctor Guntur', ...\n\n"
        "Make sure there are exactly 50 quoted queries."
    )


# --------------------------------------------------------------------------- derivation
# The vocabularies behind the six categories. Written out rather than inlined because
# `query writing.md` documents these same rules for a human, and the two must not drift.

#: Someone is looking for a PERSON, not a place or a procedure.
_PRACTITIONER = (r"doctor|doctors|dr|physician|specialist|specialists|surgeon|"
                 r"dermatologist|dermatologists|trichologist|cosmetologist")

#: An actual complaint. Deliberately excludes bare "skin" and "hair" — those appear in every
#: specialist noun in this field ("skin doctor", "hair clinic") and would make every discovery
#: query look condition-led.
_CONDITION = (r"acne|pimple|pimples|hair\s*fall|hair\s*loss|hair\s*thinning|baldness|alopecia|"
              r"dandruff|eczema|psoriasis|fungal|fungus|ringworm|rash|rashes|allergy|allergies|"
              r"pigmentation|melasma|wart|warts|mole|moles|vitiligo|scar|scars|keloid|boil|"
              r"boils|dermatitis|hyperhidrosis|sweating|infection|infections|itchy|itching|"
              r"dark\s*spots|white\s*patches|scalp|wrinkle|wrinkles|stretch\s*marks|"
              r"under\s*eye|blackheads|whiteheads|open\s*pores|sun\s*tan|hives")

#: A procedure or therapy. Rule 6: naming a treatment makes a query Condition-Based, because
#: the patient is describing what they want done, not who they want to see.
_TREATMENT = (r"treatment|treatments|therapy|surgery|transplant|grafting|removal|peel|peeling|"
              r"laser|prp|gfc|botox|filler|fillers|microdermabrasion|microneedling|"
              r"whitening|brightening|polishing|resurfacing|cryotherapy|phototherapy|"
              r"mesotherapy|glutathione|hydrafacial|dermaroller|derma\s*roller")

#: A thing you buy rather than a service you book.
_PRODUCT = (r"machine|machines|device|devices|equipment|tool|tools|"
            r"cream|creams|ointment|gel|lotion|serum|shampoo|soap|tablet|tablets|capsule|"
            r"capsules|medicine|medicines|drug|drugs|product|products|kit|kits|brand|brands|"
            r"oil|oils|supplement|supplements|spray|foam|sunscreen|moisturizer|"
            r"moisturiser|face\s*wash")
# "patch"/"patches" is deliberately absent: in dermatology it is nearly always a symptom
# ("white patches", "dry patches"), not a transdermal product, and listing it here sent
# vitiligo queries to Product-Based.

_PRICING = r"\b(?:fee|fees|cost|costs|price|prices|pricing|charges|cheap|affordable|budget)\b|₹"
_BOOKING = r"\b(?:appointment|appointments|book|booking|consult|consultation|slot|timings)\b"

_PRACTITIONER_RE = re.compile(rf"\b(?:{_PRACTITIONER})\b", re.I)
_CONDITION_RE = re.compile(rf"\b(?:{_CONDITION})\b", re.I)
_TREATMENT_RE = re.compile(rf"\b(?:{_TREATMENT})\b", re.I)
_PRODUCT_RE = re.compile(rf"\b(?:{_PRODUCT})\b", re.I)
_PRICING_RE = re.compile(_PRICING, re.I)
_BOOKING_RE = re.compile(_BOOKING, re.I)

_INTENT = {
    "Discovery": "Wants to discover the leading dermatologists in Guntur.",
    "Doctor-Based": "Wants a doctor who treats one specific condition.",
    "Condition-Based": "Is searching for treatment of a specific skin or hair condition.",
    "Product-Based": "Is looking for a machine, device or product rather than a clinic.",
    "Pricing": "Wants to know consultation fees / treatment costs.",
    "Appointment & Booking": "Is ready to book or consult a dermatologist.",
}

_CAT_WEIGHT = {
    "Discovery": 2, "Doctor-Based": 2, "Condition-Based": 1,
    "Product-Based": 0, "Pricing": 1, "Appointment & Booking": 1,
}


def derive_category(q: str) -> str:
    """Classify a query into one of the six canonical categories. Order is the rule.

    Money and booking win first: "hair transplant cost" is a price question that happens to
    name a procedure. Products come next, because a thing you buy is never a clinic visit.
    Then the distinction that matters most for this product — is the patient looking for a
    PERSON who treats their complaint (Doctor-Based) or for the PROCEDURE itself
    (Condition-Based)? Anything left is generic discovery.
    """
    ql = q.lower()
    if _PRICING_RE.search(ql):
        return "Pricing"
    if _BOOKING_RE.search(ql):
        return "Appointment & Booking"
    if _PRODUCT_RE.search(ql):
        return "Product-Based"
    has_condition = bool(_CONDITION_RE.search(ql))
    if has_condition and _PRACTITIONER_RE.search(ql):
        return "Doctor-Based"
    if has_condition or _TREATMENT_RE.search(ql):
        return "Condition-Based"
    return "Discovery"


def derive_intent(category: str) -> str:
    return _INTENT.get(category, _INTENT["Discovery"])


def derive_strength(rank: int, category: str, n_total: int) -> int:
    """Estimate a 1-10 search-strength score from rank position + small category weight."""
    base = round(10 - 9 * (rank - 1) / max(1, n_total - 1))  # rank 1 -> ~10, last -> ~1
    return max(1, min(10, base + _CAT_WEIGHT.get(category, 0) - 1))


# --------------------------------------------------------------------------- parsing
_QUOTED = re.compile(r"""['"]([^'"]*)['"]""")
_LEADING_NUM = re.compile(r"^\s*\d+[.)]\s*")
_HAS_ALNUM = re.compile(r"[A-Za-z0-9]")


def parse_pasted_queries(text: str) -> list[dict]:
    """Parse whatever the user pasted into clean query rows.

    Tolerates single/double quotes, numbering, bullets, brackets, and comma/newline separation.
    De-duplicates case-insensitively. Returns rows with all 5 fields the app expects.
    """
    if not text or not text.strip():
        return []

    candidates = _QUOTED.findall(text)
    if not candidates:  # no quotes at all -> fall back to comma/newline splitting
        candidates = re.split(r"[,\n]+", text)

    seen: set[str] = set()
    cleaned: list[str] = []
    for c in candidates:
        s = _LEADING_NUM.sub("", c).strip().strip("[]").strip().strip("'\"").strip()
        if not s or not _HAS_ALNUM.search(s):  # skip empties + stray separators like ","
            continue
        key = s.lower()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(s)

    n = len(cleaned)
    rows: list[dict] = []
    for i, q in enumerate(cleaned, start=1):
        cat = derive_category(q)
        rows.append({
            "rank": i,
            "search_query": q,
            "category": cat,
            "user_intent": derive_intent(cat),
            "search_strength_score": derive_strength(i, cat, n),
        })
    return rows


def category_breakdown(rows: list[dict]) -> dict:
    """Count rows per category (for the UI summary before saving)."""
    out: dict[str, int] = {}
    for r in rows:
        out[r["category"]] = out.get(r["category"], 0) + 1
    return out


# --------------------------------------------------------------------------- excel
_HEADERS = ["Rank", "Search Query", "Category", "User Intent", "Search Strength Score"]
_KEYS = ["rank", "search_query", "category", "user_intent", "search_strength_score"]


def save_queries_xlsx(rows: list[dict], path: str | None = None) -> str:
    """Write rows to a formatted .xlsx (bold header, blue fill, frozen top row, auto width)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = path or config.QUERIES_XLSX
    Path(path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Queries"

    header_fill = PatternFill("solid", fgColor="DDEEFF")
    header_font = Font(bold=True)
    for col, name in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, key in enumerate(_KEYS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(key))

    # auto column width (capped)
    for c_idx, key in enumerate(_KEYS, start=1):
        longest = len(_HEADERS[c_idx - 1])
        for row in rows:
            longest = max(longest, len(str(row.get(key, ""))))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(longest + 2, 60)

    ws.freeze_panes = "A2"
    wb.save(path)
    return path
