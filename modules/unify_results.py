"""Task 2 — unify the Maps dataset and the screenshot google-search dataset into one per-clinic view.

Built only after both independent datasets exist. The Maps export is appearance-level (750 rows), so
the Maps side is aggregated to clinic level (clean dedup keys) before joining — a literal merge of the
two xlsx sheets on keyword-stuffed names would be wrong. Emits data/unified_results.xlsx and exposes
the per-clinic web signal that feeds the 40% web term (vulnerability.web_relevance_vuln).
"""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd

try:
    import config
except ModuleNotFoundError:  # pragma: no cover
    import importlib
    config = importlib.import_module("config")

from modules.maps_collector import dedup_key

# Web-signal columns merged onto each clinic (defaults applied when a clinic never appears on a SERP).
_WEB_DEFAULTS = {
    "web_appearances": 0, "web_owned_appearances": 0, "web_borrowed_appearances": 0,
    "web_best_position": None, "has_own_site": False, "in_places_count": 0,
    "sponsored_count": 0, "ai_overview_count": 0,
}


def _clinic_key(row) -> str:
    return dedup_key(row.get("place_url", "")) or str(row.get("name") or "").strip().lower()


def unify(maps_df: pd.DataFrame, web_by_clinic: dict) -> pd.DataFrame:
    """Join the clinic-level Maps view with the per-clinic web signal (keyed by dedup key).

    `web_data` is True for every clinic when any web data exists (an absent clinic simply never
    appeared = fully invisible, which is real signal — not "no data"); False everywhere when web was
    never collected, so the score stays Maps-only.
    """
    web_exists = bool(web_by_clinic)
    rows = []
    for _, r in maps_df.iterrows():
        key = _clinic_key(r)
        w = web_by_clinic.get(key) or {}
        row = r.to_dict()
        row["clinic_key"] = key
        row["web_data"] = bool(w.get("web_data", web_exists))
        for field, default in _WEB_DEFAULTS.items():
            val = w.get(field, default)
            if field == "web_best_position":
                row[field] = val
            elif field == "has_own_site":
                row[field] = bool(val)
            else:
                row[field] = int(val or 0)
        rows.append(row)
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------- export
_HEADERS = ["Clinic", "Rating", "Reviews", "Website", "Maps Appearances", "Web Appearances",
            "Owned (own-site/ads)", "Borrowed (aggregators)", "Best Web Pos", "Has Own Site",
            "In Places", "Sponsored", "AI Overview", "Opportunity Score", "Label"]


def save_unified_xlsx(df: pd.DataFrame, path: str | None = None) -> str:
    """Write the unified per-clinic view (Maps signals ‖ web-presence signals [+ score if scored])."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = path or config.UNIFIED_XLSX
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sh = wb.active
    sh.title = "Unified"
    bold, fill = Font(bold=True), PatternFill("solid", fgColor="DDEEFF")
    for col, name in enumerate(_HEADERS, start=1):
        cell = sh.cell(row=1, column=col, value=name)
        cell.font, cell.fill = bold, fill

    for i, (_, r) in enumerate(df.iterrows(), start=2):
        sh.cell(row=i, column=1, value=r.get("name"))
        sh.cell(row=i, column=2, value=r.get("rating"))
        sh.cell(row=i, column=3, value=r.get("user_ratings_total"))
        sh.cell(row=i, column=4, value=(r.get("website") or ""))
        sh.cell(row=i, column=5, value=r.get("appearances"))
        sh.cell(row=i, column=6, value=r.get("web_appearances"))
        sh.cell(row=i, column=7, value=r.get("web_owned_appearances"))
        sh.cell(row=i, column=8, value=r.get("web_borrowed_appearances"))
        sh.cell(row=i, column=9, value=r.get("web_best_position"))
        sh.cell(row=i, column=10, value="yes" if r.get("has_own_site") else "")
        sh.cell(row=i, column=11, value=r.get("in_places_count"))
        sh.cell(row=i, column=12, value=r.get("sponsored_count"))
        sh.cell(row=i, column=13, value=r.get("ai_overview_count"))
        sh.cell(row=i, column=14, value=r.get("vulnerability_score"))
        sh.cell(row=i, column=15, value=r.get("vulnerability_label"))

    for idx, w in enumerate([30, 8, 8, 34, 16, 15, 18, 20, 12, 12, 10, 10, 10, 16, 10], start=1):
        sh.column_dimensions[get_column_letter(idx)].width = w
    sh.freeze_panes = "A2"
    wb.save(path)
    return path


# --------------------------------------------------------------------------- real-run orchestration
def web_signal_by_clinic(web_screens: dict | None = None, clinics: list[dict] | None = None) -> dict:
    """Per-clinic web signal from the screenshot dataset (the input to the 40% web term)."""
    from modules import web_screens as ws_mod
    if web_screens is None:
        web_screens = json.loads(Path(config.WEB_SCREENS_CACHE).read_text(encoding="utf-8"))
    if clinics is None:
        clinics = ws_mod.clinics_from_maps()
    return ws_mod.aggregate_web_by_clinic(web_screens, clinics)


def build_unified() -> pd.DataFrame:
    """Full Task-2 run: aggregate Maps -> attach web signal -> score (presence-weighted) -> xlsx."""
    from modules import storage, web_screens as ws_mod
    from modules.vulnerability import aggregate_clinics, score_clinics

    rows = storage.load_rows(storage.RESULTS_JSON) or []
    maps_df = aggregate_clinics(rows)
    screens = json.loads(Path(config.WEB_SCREENS_CACHE).read_text(encoding="utf-8"))
    clinics = [{"name": r.get("name"), "website": r.get("website"), "place_url": r.get("place_url")}
               for _, r in maps_df.iterrows()]
    web_by = ws_mod.aggregate_web_by_clinic(screens, clinics)
    unified = unify(maps_df, web_by)
    scored = score_clinics(unified)            # web_* columns present -> presence-weighted 40% blend
    save_unified_xlsx(scored)
    return scored


if __name__ == "__main__":  # pragma: no cover - manual run
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    df = build_unified()
    print(f"unified {len(df)} clinics -> {config.UNIFIED_XLSX}")
