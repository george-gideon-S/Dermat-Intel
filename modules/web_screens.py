"""Step 3b — structured google-search dataset from the manual SERP screenshots (free, no API key).

Live Google web-search scraping is blocked, so we captured full-page SERP screenshots, sliced them
into legible tiles (screenshot_slicer), and extracted each result block via Claude vision into
per-part JSON. This module is the deterministic half: reconcile parts -> map results to the 34 Maps
clinics -> per-clinic OWNED-vs-BORROWED web visibility that feeds the 40% web term.

Independent of the Maps dataset (Task-1 deliverable). Matching reuses the proven, tested helpers from
web_collector so the two web paths classify clinics identically.

Key scoring decision: OWNED web visibility = own-site organic ranking OR a paid ad. The Places (local
pack) is Google **Maps** data re-surfaced on the SERP, so it is captured/reported but is NOT counted as
owned web visibility — otherwise the 40% web term would just echo the 60% Maps term.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

try:
    import config
except ModuleNotFoundError:  # pragma: no cover
    import importlib
    config = importlib.import_module("config")

from modules.maps_collector import dedup_key
from modules.web_collector import _name_tokens, _result_matches_clinic, domain_of

# Result-page platforms that constitute BORROWED visibility (the clinic is found only because a
# third-party directory / social profile lists it — it owns none of that real estate).
AGGREGATOR_PLATFORMS = {"practo", "justdial", "lybrate", "skedoc", "sulekha", "drlogy",
                        "apollo247", "bajajfinservhealth"}
SOCIAL_PLATFORMS = {"instagram", "facebook", "youtube"}
BORROWED_PLATFORMS = AGGREGATOR_PLATFORMS | SOCIAL_PLATFORMS | {"traya"}
SPONSORED_TYPES = {"sponsored_top", "sponsored_mid"}
VALID_BLOCK_TYPES = {"sponsored_top", "sponsored_mid", "places", "organic", "ai_overview"}


def _normalize_block(b: dict) -> dict:
    """Repair vision blocks whose block_type holds a platform string (fields occasionally swapped).

    An unrecognised block_type means a plain organic result whose platform leaked into block_type —
    coerce block_type to 'organic' and recover the platform when it wasn't separately captured.
    """
    bt = b.get("block_type")
    if bt in VALID_BLOCK_TYPES:
        return b
    out = dict(b)
    if (not out.get("platform")) or out.get("platform") == "other":
        if bt and bt != "other":
            out["platform"] = bt
    out["block_type"] = "organic"
    return out


# --------------------------------------------------------------------------- query reconciliation
def _norm(s: str) -> str:
    """Lowercase, strip punctuation to spaces, collapse whitespace — for query matching."""
    s = re.sub(r"[^a-z0-9]+", " ", (s or "").lower())
    return re.sub(r"\s+", " ", s).strip()


def match_query(search_box_text: str, query_rows: list[dict]) -> dict:
    """Map a screenshot's search-box text to a canonical query. Authoritative over capture order.

    Exact normalized match first; else best token-Jaccard ≥ 0.6 (tolerates minor vision noise); else
    unmatched (rank None) — surfaced in the reconciliation report rather than guessed.
    """
    nt = _norm(search_box_text)
    for qr in query_rows:
        if _norm(qr.get("search_query", "")) == nt and nt:
            return {"rank": qr.get("rank"), "search_query": qr.get("search_query"),
                    "match_confidence": "exact"}
    tt = set(nt.split())
    best, best_j = None, 0.0
    for qr in query_rows:
        qs = set(_norm(qr.get("search_query", "")).split())
        union = tt | qs
        j = (len(tt & qs) / len(union)) if union else 0.0
        if j > best_j:
            best, best_j = qr, j
    if best and best_j >= 0.6:
        return {"rank": best.get("rank"), "search_query": best.get("search_query"),
                "match_confidence": "fuzzy"}
    return {"rank": None, "search_query": None, "match_confidence": "unmatched"}


def reconcile(parts: list[dict], manifest: dict, query_rows: list[dict]) -> dict:
    """Merge the per-agent part files into one dataset, mapping each screenshot to its query.

    Reconciles the 78-vs-80 gap: queries with no matching screenshot are reported in
    meta.unmatched_queries (a real signal — those SERPs were never captured).
    """
    idx_to_shot = {s.get("index"): s.get("screenshot") for s in manifest.get("screenshots", [])}
    entries = [q for part in parts for q in part.get("queries", [])]
    entries.sort(key=lambda q: q.get("index", 1_000_000))

    out_queries, matched = [], set()
    for q in entries:
        m = match_query(q.get("search_box_text", ""), query_rows)
        if m["rank"] is not None:
            matched.add(m["search_query"])
        out_queries.append({
            "index": q.get("index"),
            "screenshot": idx_to_shot.get(q.get("index")),
            "rank": m["rank"],
            "search_query": m["search_query"],
            "search_box_text": q.get("search_box_text", ""),
            "match_confidence": m["match_confidence"],
            "readable": q.get("readable", True),
            "blocks": [_normalize_block(b) for b in q.get("blocks", [])],
        })
    unmatched = [qr.get("search_query") for qr in query_rows if qr.get("search_query") not in matched]
    return {
        "meta": {
            "num_screenshots": manifest.get("num_screenshots", len(entries)),
            "num_queries_expected": manifest.get("num_queries_expected", len(query_rows)),
            "unmatched_queries": unmatched,
            "tile_h": manifest.get("tile_h"),
            "tile_overlap": manifest.get("overlap"),
        },
        "queries": out_queries,
    }


# --------------------------------------------------------------------------- clinic mapping
def prepare_clinics(clinics: list[dict]) -> list[dict]:
    """Pre-compute the match identity (key, distinctive name tokens, own-site domain) per clinic."""
    out = []
    for c in clinics:
        key = dedup_key(c.get("place_url", "")) or str(c.get("name") or "").strip().lower()
        if not key:
            continue
        out.append({"key": key, "tokens": _name_tokens(c.get("name", "")),
                    "domain": domain_of(c.get("website", "")), "name": c.get("name")})
    return out


def _is_own_site(block: dict, clinic_domain: str) -> bool:
    """True if this result lives on the clinic's OWN web property (domain match, or a clinic_site)."""
    bdom = domain_of(block.get("domain") or block.get("url") or "")
    if clinic_domain and bdom and (
        bdom == clinic_domain or bdom.endswith("." + clinic_domain) or clinic_domain.endswith("." + bdom)
    ):
        return True
    return block.get("platform") == "clinic_site"


def map_block(block: dict, prepared: list[dict]) -> tuple[str | None, bool]:
    """Resolve a result block to a clinic key (or None) + whether it sits on the clinic's own site."""
    result = {"title": block.get("title", ""), "url": block.get("url", ""),
              "domain": block.get("domain", "")}
    for c in prepared:
        if _result_matches_clinic(result, c["tokens"], c["domain"]):
            return c["key"], _is_own_site(block, c["domain"])
    return None, False


# --------------------------------------------------------------------------- per-clinic aggregation
def aggregate_web_by_clinic(web_screens: dict, clinics: list[dict]) -> dict:
    """Per-clinic web-visibility signal across all captured SERPs.

    Returns {key: {web_data, web_appearances, web_owned_appearances, web_borrowed_appearances,
    web_best_position, has_own_site, in_places_count, sponsored_count, ai_overview_count}}.
    A clinic appears in a query at most once per category. OWNED = own-site organic OR a paid ad;
    BORROWED = appears only via an aggregator/social listing; Places presence counts toward total
    appearances only (it is Maps re-surfaced, not owned web visibility).
    """
    prepared = prepare_clinics(clinics)
    out = {c["key"]: {"web_data": True, "web_appearances": 0, "web_owned_appearances": 0,
                      "web_borrowed_appearances": 0, "web_best_position": None,
                      "has_own_site": False, "in_places_count": 0, "sponsored_count": 0,
                      "ai_overview_count": 0, "platforms": set()} for c in prepared}

    for q in web_screens.get("queries", []):
        per: dict[str, list[tuple[dict, bool]]] = {}
        for b in q.get("blocks", []):
            key, own = map_block(b, prepared)
            if key is not None:
                per.setdefault(key, []).append((b, own))

        for key, items in per.items():
            rec = out.get(key)
            if rec is None:
                continue
            rec["web_appearances"] += 1
            positions = [b.get("position") for b, _ in items if b.get("position") is not None]
            if positions:
                bp = min(positions)
                rec["web_best_position"] = bp if rec["web_best_position"] is None else min(rec["web_best_position"], bp)

            owned = any(b.get("block_type") in SPONSORED_TYPES or
                        (b.get("block_type") == "organic" and own) for b, own in items)
            organic_own = any(b.get("block_type") == "organic" and own for b, own in items)
            borrowed = (not owned) and any(
                b.get("block_type") == "organic" and b.get("platform") in BORROWED_PLATFORMS
                for b, _ in items)

            if owned:
                rec["web_owned_appearances"] += 1
            if borrowed:
                rec["web_borrowed_appearances"] += 1
            if organic_own:
                rec["has_own_site"] = True
            if any(b.get("block_type") == "places" for b, _ in items):
                rec["in_places_count"] += 1
            if any(b.get("block_type") in SPONSORED_TYPES for b, _ in items):
                rec["sponsored_count"] += 1
            if any(b.get("block_type") == "ai_overview" for b, _ in items):
                rec["ai_overview_count"] += 1
            for b, _ in items:                       # third-party real-estate carrying the clinic
                if b.get("platform") in BORROWED_PLATFORMS:
                    rec["platforms"].add(b["platform"])

    for rec in out.values():
        rec["platforms"] = sorted(rec["platforms"])
    return out


# --------------------------------------------------------------------------- export (Task-1 xlsx)
_HEADERS = ["Query", "Rank", "Block Type", "Platform", "Title", "Domain", "URL",
            "Position", "Rating", "Reviews", "Mapped Clinic", "Own Site?"]


def to_rows(web_screens: dict, clinics: list[dict]) -> list[dict]:
    """One row per result block, with its mapped clinic + own-site flag."""
    prepared = prepare_clinics(clinics)
    name_by_key = {c["key"]: c["name"] for c in prepared}
    rows = []
    for q in web_screens.get("queries", []):
        for b in q.get("blocks", []):
            key, own = map_block(b, prepared)
            rows.append({
                "query": q.get("search_query") or q.get("search_box_text"),
                "rank": q.get("rank"),
                "block_type": b.get("block_type"),
                "platform": b.get("platform"),
                "title": b.get("title"),
                "domain": b.get("domain", ""),
                "url": b.get("url", ""),
                "position": b.get("position"),
                "rating": b.get("rating"),
                "reviews": b.get("reviews"),
                "mapped_clinic": name_by_key.get(key, "") if key else "",
                "mapped_key": key or "",
                "is_own_site": bool(own) if key else False,
            })
    return rows


def save_search_xlsx(rows: list[dict], path: str | None = None) -> str:
    """Write the google-search dataset (one row per result block) to xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = path or config.SEARCH_RESULTS_XLSX
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sh = wb.active
    sh.title = "Google Search"
    bold, fill = Font(bold=True), PatternFill("solid", fgColor="DDEEFF")
    for col, name in enumerate(_HEADERS, start=1):
        cell = sh.cell(row=1, column=col, value=name)
        cell.font, cell.fill = bold, fill
    for i, r in enumerate(rows, start=2):
        sh.cell(row=i, column=1, value=r.get("query"))
        sh.cell(row=i, column=2, value=r.get("rank"))
        sh.cell(row=i, column=3, value=r.get("block_type"))
        sh.cell(row=i, column=4, value=r.get("platform"))
        sh.cell(row=i, column=5, value=r.get("title"))
        sh.cell(row=i, column=6, value=r.get("domain"))
        sh.cell(row=i, column=7, value=r.get("url"))
        sh.cell(row=i, column=8, value=r.get("position"))
        sh.cell(row=i, column=9, value=r.get("rating"))
        sh.cell(row=i, column=10, value=r.get("reviews"))
        sh.cell(row=i, column=11, value=r.get("mapped_clinic"))
        sh.cell(row=i, column=12, value="yes" if r.get("is_own_site") else "")
    for idx, w in enumerate([34, 6, 14, 12, 44, 26, 30, 8, 7, 8, 26, 9], start=1):
        sh.column_dimensions[get_column_letter(idx)].width = w
    sh.freeze_panes = "A2"
    wb.save(path)
    return path


# --------------------------------------------------------------------------- real-run orchestration
def load_parts(cache_dir: str | None = None) -> list[dict]:
    cache_dir = Path(cache_dir or config.CACHE_DIR)
    parts = []
    for p in sorted(cache_dir.glob("web_screens_part_*.json")):
        try:
            with open(p, encoding="utf-8") as fh:
                parts.append(json.load(fh))
        except (json.JSONDecodeError, OSError):
            continue
    return parts


def save_web_screens(screens: dict, path: str | None = None) -> str:
    path = path or config.WEB_SCREENS_CACHE
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(screens, fh, ensure_ascii=False, indent=2)
    return path


def load_web_screens(path: str | None = None) -> dict:
    """Load the persisted screenshot dataset; {} if it doesn't exist yet."""
    path = path or config.WEB_SCREENS_CACHE
    try:
        with open(path, encoding="utf-8") as fh:
            return json.load(fh)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def clinics_from_maps() -> list[dict]:
    """The 34 Maps clinics as lightweight match records (name/website/place_url)."""
    from modules import storage
    from modules.vulnerability import aggregate_clinics
    rows = storage.load_rows(storage.RESULTS_JSON) or []
    agg = aggregate_clinics(rows)
    return [{"name": r.get("name"), "website": r.get("website"), "place_url": r.get("place_url")}
            for _, r in agg.iterrows()]


def build_dataset() -> dict:
    """Merge parts -> reconcile -> persist web_screens.json -> map clinics -> write search xlsx."""
    from modules import storage
    parts = load_parts()
    manifest = json.loads((Path(config.WEB_TILES_DIR) / "manifest.json").read_text(encoding="utf-8"))
    qrows = storage.load_rows(storage.QUERIES_JSON) or []
    screens = reconcile(parts, manifest, qrows)
    save_web_screens(screens)
    clinics = clinics_from_maps()
    save_search_xlsx(to_rows(screens, clinics))
    screens["_web_by_clinic"] = aggregate_web_by_clinic(screens, clinics)
    return screens


if __name__ == "__main__":  # pragma: no cover - manual run
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    s = build_dataset()
    print(f"queries: {len(s['queries'])}  unmatched: {s['meta']['unmatched_queries']}")
