import openpyxl

from modules import maps_collector as mc
from modules.maps_collector import (
    RESULT_COLUMNS,
    make_mock_results,
    dedup_key,
    parse_listing,
    save_results_xlsx,
)

QROWS = [{
    "rank": 1, "search_query": "best dermatologist in Guntur", "category": "Discovery",
    "user_intent": "x", "search_strength_score": 10,
}]


# ---------------------------------------------------------------- pure helpers
def test_mock_results_shape():
    rows = make_mock_results(QROWS, per_query=15)
    assert len(rows) == 15
    assert set(RESULT_COLUMNS).issubset(rows[0].keys())
    assert rows[0]["source_query"] == "best dermatologist in Guntur"
    assert rows[0]["result_position"] == 1


def test_dedup_key_from_cid_url():
    u = "https://www.google.com/maps/place/?q=place_id:ChIJabc&cid=12345"
    assert dedup_key(u) == "12345"


def test_dedup_key_falls_back_to_place_id():
    u = "https://www.google.com/maps/place/?q=place_id:ChIJ_abc-123"
    assert dedup_key(u) == "ChIJ_abc-123"


def test_parse_listing_maps_fields():
    raw = {
        "name": "Skin Clinic", "rating": 4.6, "reviews": 120, "address": "MG Road, Guntur",
        "phone": "+91 90000 11111", "website": "http://x.com", "types": "Dermatologist",
        "lat": 16.30, "lng": 80.43, "url": "https://www.google.com/maps/place/?cid=99",
        "closed": False,
    }
    row = parse_listing(raw, QROWS[0], position=3)
    assert row["name"] == "Skin Clinic"
    assert row["rating"] == 4.6
    assert row["user_ratings_total"] == 120
    assert row["result_position"] == 3
    assert row["source_query_rank"] == 1
    assert row["business_status"] == "OPERATIONAL"
    assert row["status"] == "OK"
    assert row["place_id"] == "99"


# ---------------------------------------------------------------- network-mocked
def test_geocode_osm_parses(monkeypatch):
    class R:
        status_code = 200

        def json(self):
            return [{"lat": "16.31", "lon": "80.44"}]

    monkeypatch.setattr(mc.requests, "get", lambda *a, **k: R())
    assert mc.geocode_osm("MG Road, Guntur") == (16.31, 80.44)


def test_geocode_osm_handles_empty(monkeypatch):
    class R:
        status_code = 200

        def json(self):
            return []

    monkeypatch.setattr(mc.requests, "get", lambda *a, **k: R())
    assert mc.geocode_osm("nowhere") is None


def test_collect_mock_does_not_touch_network(monkeypatch):
    def boom(*a, **k):
        raise AssertionError("network must not be used in mock mode")

    monkeypatch.setattr(mc, "_run_browser", boom)
    rows = mc.collect(QROWS, mock=True)
    assert len(rows) == 15
    assert rows[0]["status"] == "OK"


def test_run_browser_wrapper_returns_and_reraises(monkeypatch):
    # The wrapper runs _run_browser_impl in a worker thread (Streamlit asyncio fix).
    # It must return the impl's result and propagate the impl's exceptions.
    monkeypatch.setattr(mc, "_run_browser_impl", lambda *a, **k: ["ok"])
    assert mc._run_browser("q", 3) == ["ok"]

    def boom(*a, **k):
        raise RuntimeError("scrape blew up")

    monkeypatch.setattr(mc, "_run_browser_impl", boom)
    import pytest
    with pytest.raises(RuntimeError):
        mc._run_browser("q", 3)


# ---------------------------------------------------------------- excel
def test_save_results_xlsx(tmp_path):
    rows = make_mock_results(QROWS, per_query=15)
    out = save_results_xlsx(rows, str(tmp_path / "r.xlsx"))
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.freeze_panes == "A2"
    assert ws.max_column == len(RESULT_COLUMNS)
    assert ws.max_row == len(rows) + 1
