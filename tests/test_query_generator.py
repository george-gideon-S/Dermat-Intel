import openpyxl
import pytest

from modules.query_generator import (
    build_ai_prompt,
    parse_pasted_queries,
    derive_category,
    save_queries_xlsx,
)


# ---------------------------------------------------------------- build_ai_prompt
def test_prompt_demands_50_and_format():
    p = build_ai_prompt()
    assert "exactly 50" in p.lower()
    assert "comma-separated" in p.lower()
    assert "'best dermatologist in Guntur'" in p  # format example present
    assert "Guntur" in p


# ---------------------------------------------------------------- parse
def test_parse_simple_single_quotes():
    rows = parse_pasted_queries("'best dermatologist in Guntur', 'skin specialist near me'")
    assert [r["search_query"] for r in rows] == [
        "best dermatologist in Guntur",
        "skin specialist near me",
    ]
    assert rows[0]["rank"] == 1 and rows[1]["rank"] == 2


def test_parse_double_quotes_newlines_numbering_brackets():
    raw = '[\n 1. "acne treatment doctor Guntur",\n 2. "dermatologist fees Guntur"\n]'
    rows = parse_pasted_queries(raw)
    assert [r["search_query"] for r in rows] == [
        "acne treatment doctor Guntur",
        "dermatologist fees Guntur",
    ]


def test_parse_dedup_case_insensitive_and_empties():
    rows = parse_pasted_queries("'a clinic', 'A Clinic', '', '  '")
    assert [r["search_query"] for r in rows] == ["a clinic"]


def test_parse_no_quotes_falls_back_to_commas():
    rows = parse_pasted_queries("best dermatologist Guntur, acne treatment Guntur")
    assert [r["search_query"] for r in rows] == [
        "best dermatologist Guntur",
        "acne treatment Guntur",
    ]


def test_parse_empty_returns_empty():
    assert parse_pasted_queries("") == []
    assert parse_pasted_queries("   \n  ") == []


@pytest.mark.parametrize("q,cat", [
    # Rule 8 — a condition PLUS a practitioner means the patient wants a person.
    ("acne treatment doctor Guntur", "Doctor-Based"),
    ("hair fall specialist Guntur", "Doctor-Based"),
    ("best hair fall doctor in Guntur", "Doctor-Based"),
    ("psoriasis specialist in Guntur", "Doctor-Based"),
    # ...but a practitioner with no condition is still plain discovery.
    ("best dermatologist in Guntur", "Discovery"),
    ("lady skin doctor in Guntur", "Discovery"),
    # Rule 6 — naming a treatment is condition intent, never discovery.
    ("laser hair removal in Guntur", "Condition-Based"),
    ("hair transplant in Guntur", "Condition-Based"),
    ("scalp treatment in Guntur", "Condition-Based"),
    ("acne treatment in Guntur", "Condition-Based"),
    # Rule 7 — a thing you buy, not a visit you book.
    ("best laser machine for hair removal in Guntur", "Product-Based"),
    ("acne cream in Guntur", "Product-Based"),
    # ...but a symptom that merely sounds like a product is not one.
    ("white patches treatment in Guntur", "Condition-Based"),
    # Rule 11 — money and booking outrank the treatment they name.
    ("dermatologist fees in Guntur", "Pricing"),
    ("hair transplant cost in Guntur", "Pricing"),
    ("book dermatologist appointment Guntur", "Appointment & Booking"),
    # Rule 9 — trust/social proof is discovery wearing a different hat.
    ("best rated skin doctor reviews Guntur", "Discovery"),
    ("dermatologist reviews Guntur", "Discovery"),
    # Rule 10 — comparison is retired; these are discovery.
    ("dermatologist vs cosmetologist Guntur", "Discovery"),
    ("best skin clinic or dermatologist in Guntur", "Discovery"),
])
def test_derive_category(q, cat):
    assert derive_category(q) == cat


def test_the_retired_categories_can_never_be_produced():
    """Comparison and Trust & Social Proof are gone — nothing may still emit them."""
    import config
    retired = {"Comparison", "Trust & Social Proof", "Near Me / Local"}
    assert not (set(config.CATEGORIES) & retired)
    probes = ["dermatologist vs cosmetologist Guntur", "best rated dermatologist in Guntur",
              "skin specialist reviews Guntur", "compare skin clinics in Guntur",
              "skin specialist near me", "dermatologist ratings Guntur"]
    assert not ({derive_category(q) for q in probes} & retired)


def test_every_derived_category_is_a_declared_one():
    """A category the config does not declare has no colour and no place in the report."""
    import config
    probes = ["dermatologist in Guntur", "acne doctor in Guntur", "acne treatment in Guntur",
              "acne cream in Guntur", "dermatologist fees in Guntur",
              "book appointment dermatologist Guntur", "xyzzy in Guntur"]
    for q in probes:
        assert derive_category(q) in config.CATEGORIES, q


def test_all_rows_have_five_fields():
    rows = parse_pasted_queries("'best dermatologist in Guntur'")
    assert set(rows[0]) == {
        "rank", "search_query", "category", "user_intent", "search_strength_score",
    }
    assert 1 <= rows[0]["search_strength_score"] <= 10


# ---------------------------------------------------------------- save_queries_xlsx
def test_save_xlsx(tmp_path):
    rows = parse_pasted_queries("'best dermatologist in Guntur', 'acne treatment Guntur'")
    out = save_queries_xlsx(rows, str(tmp_path / "q.xlsx"))
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert [c.value for c in ws[1]] == [
        "Rank", "Search Query", "Category", "User Intent", "Search Strength Score",
    ]
    assert ws[1][0].font.bold is True
    assert ws.freeze_panes == "A2"
    assert ws.max_row == 3  # header + 2 rows
