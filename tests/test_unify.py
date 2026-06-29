"""Tests for modules.unify_results — join the Maps clinic view with the screenshot web signal.

Pure join logic on synthetic data; no real datasets touched.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import pandas as pd

from modules import unify_results as uni


def _maps_df():
    return pd.DataFrame([
        {"name": "Dr Sowmya Skin Hair Laser Clinic", "website": "https://drsowmyaskinclinics.com/",
         "place_url": "https://www.google.com/maps/place/?cid=1", "rating": 4.8,
         "user_ratings_total": 519, "appearances": 20},
        {"name": "Skin Perfect Clinic", "website": "",
         "place_url": "https://www.google.com/maps/place/?cid=3", "rating": 4.9,
         "user_ratings_total": 611, "appearances": 25},
    ])


def _web():
    return {
        "1": {"web_data": True, "web_appearances": 2, "web_owned_appearances": 1,
              "web_borrowed_appearances": 1, "web_best_position": 5, "has_own_site": True,
              "in_places_count": 0, "sponsored_count": 0, "ai_overview_count": 0},
        "3": {"web_data": True, "web_appearances": 1, "web_owned_appearances": 0,
              "web_borrowed_appearances": 0, "web_best_position": 1, "has_own_site": False,
              "in_places_count": 1, "sponsored_count": 0, "ai_overview_count": 0},
    }


def test_unify_merges_web_signal_onto_each_clinic():
    out = uni.unify(_maps_df(), _web())
    assert len(out) == 2
    sow = out[out["name"].str.contains("Sowmya")].iloc[0]
    assert sow["web_owned_appearances"] == 1 and bool(sow["has_own_site"]) is True
    assert sow["rating"] == 4.8 and sow["appearances"] == 20      # Maps columns preserved
    perfect = out[out["name"].str.contains("Perfect")].iloc[0]
    assert perfect["web_owned_appearances"] == 0 and perfect["in_places_count"] == 1
    assert bool(out["web_data"].all()) is True                    # activates the 40% web term


def test_unify_absent_clinic_gets_zero_presence_but_web_data_true():
    # a clinic that never appears in any SERP is fully invisible — zeros, but web data exists
    maps = pd.DataFrame([{"name": "Ghost Clinic", "website": "",
                          "place_url": "https://www.google.com/maps/place/?cid=9",
                          "rating": 5.0, "user_ratings_total": 3, "appearances": 1}])
    out = uni.unify(maps, _web())                                 # cid=9 not in web signal
    row = out.iloc[0]
    assert row["web_appearances"] == 0 and row["web_owned_appearances"] == 0
    assert bool(row["web_data"]) is True
    assert row["clinic_key"] == "9"


def test_unify_no_web_data_marks_web_data_false():
    out = uni.unify(_maps_df(), {})                               # web never collected
    assert bool(out["web_data"].any()) is False                  # -> score stays Maps-only


def test_save_unified_xlsx(tmp_path):
    out = uni.unify(_maps_df(), _web())
    p = uni.save_unified_xlsx(out, str(tmp_path / "unified.xlsx"))
    ws = openpyxl.load_workbook(p).active
    assert ws.max_row == len(out) + 1            # header + 2 clinics
    assert ws["A1"].value is not None
